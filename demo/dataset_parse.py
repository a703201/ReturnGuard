"""ReturnGuard · 上传数据集文件解析（网页端「文件导入」用）

把卖家上传的真实数据集文件（.xlsx / .csv）解析为 ReturnGuard 案件 dict 列表，
供 importer.import_file 批量落库（带按 case_id 去重）。

与 convert_datasets.py 的区别：
- 解析「上传的字节流」，而非 Dataset/ 固定路径；
- 全量导入（不抽样），保留原始退货日期（不平移、不注入爆发），诚实回流；
- 为每个案件生成稳定 case_id（AMZ-/UCI-/TL-/RG-），使重复导入可幂等去重；
- 复用 convert_datasets.py 的品类/缺陷/供应商映射（单一事实来源），导入失败回退内联兜底。

支持类型（按表头特征识别）：
- .xlsx → Amazon 退货（含 return_reason）/ UCI Online Retail（退货=负数量或 C 开头单）
- .csv  → TheLook 订单（returned_at 非空）或 RG 格式（含 sku/amount 等列）
"""
from __future__ import annotations

import csv
import io
import logging
import re

try:
    import openpyxl
except ImportError:  # noqa: BLE001
    openpyxl = None

logger = logging.getLogger("returnguard.dataset_parse")

# ---- 复用 convert_datasets 的受控词表（单一事实来源），失败回退内联 ----
try:
    from convert_datasets import (  # type: ignore
        AMAZON_CAT_MAP, AMAZON_REASON_DEFECT, REMORSE_REASONS, QUALITY_REASONS,
        CATEGORY_KEYWORDS, THELOOK_CAT_MAP, _supplier_for, _map_category,
        _clamp, SAME_ITEM_THRESHOLD, _to_date, SUPPLIERS,
    )
except Exception:  # noqa: BLE001
    logger.warning("无法从 convert_datasets 导入映射词表，使用内联兜底", exc_info=True)
    SAME_ITEM_THRESHOLD = 0.82
    SUPPLIERS = {
        "S1": "鼎峰精密", "S2": "云仓优选", "S3": "鑫源电子(劣)", "S4": "通达包装弱",
        "S5": "联创供货", "S6": "海贸乱发(劣)", "S7": "锐捷制造", "S8": "万通杂货",
    }
    AMAZON_REASON_DEFECT = {
        "Wrong item": "货不对板", "Size/fit issue": "货不对板", "Changed mind": "无明显瑕疵",
        "Duplicate order": "无明显瑕疵", "Missing parts": "商品缺件", "Quality issue": "功能故障",
        "Damaged": "外包装破损", "Not as described": "货不对板", "Late delivery": "无明显瑕疵",
    }
    REMORSE_REASONS = {"Changed mind", "Duplicate order", "Late delivery"}
    QUALITY_REASONS = {"Quality issue", "Damaged", "Missing parts", "Wrong item", "Not as described", "Size/fit issue"}
    CATEGORY_KEYWORDS = [
        ("3C数码", ["phone", "charger", "headset", "headphone", "speaker", "cable", "usb",
                    "battery", "lamp", "light", "plug", "adapter", "camera", "wire", "electronic",
                    "earphone", "mouse", "keyboard", "tablet", "computer", "watch", "tech"]),
        ("饰品配件", ["necklace", "ring", "earring", "bracelet", "pendant", "brooch", "chain",
                    "jewel", "hair", "clip", "bead", "crystal", "gem", "tiara", "necklet",
                    "accessory", "scarf", "glove", "sock", "hat", "beauty", "cosmetic",
                    "toy", "game", "book", "swim"]),
        ("小家电", ["heater", "bottle", "iron", "toaster", "blender", "kettle", "fryer",
                   "mixer", "purifier", "humidifier", "fan", "vacuum", "cooker", "warmer",
                   "holder", "cup", "mug", "lantern", "heart", "door", "sign",
                   "clock", "frame", "vase", "cushion", "rug", "bowl", "plate", "home",
                   "kitchen", "house", "decor", "appliance"]),
        ("服饰鞋包", ["bag", "shoe", "boot", "dress", "shirt", "trouser", "coat", "cardigan",
                    "apron", "jumper", "top", "sweater", "pant", "wallet", "purse", "cap",
                    "clothing", "apparel", "footwear", "sport", "wear"]),
    ]
    AMAZON_CAT_MAP = {
        "Electronics": "3C数码", "Clothing": "服饰鞋包", "Sports": "服饰鞋包",
        "Beauty": "饰品配件", "Toys": "饰品配件", "Books": "饰品配件", "Home": "小家电",
    }
    THELOOK_CAT_MAP = {
        "Electronics": "3C数码", "Footwear": "服饰鞋包", "Apparel": "服饰鞋包",
        "Swim": "服饰鞋包", "Athletic": "服饰鞋包", "Accessories": "饰品配件",
        "Beauty": "饰品配件", "Intimate": "服饰鞋包", "Home": "小家电", "Kitchen": "小家电",
        "Toy": "饰品配件", "Books": "饰品配件", "Furniture": "小家电",
    }
    QUALITY_DEFECTS = {"功能故障", "货不对板", "商品缺件", "外包装破损", "污渍划痕", "色差明显", "使用痕迹"}
    GOOD_SUPPLIERS = ["S1", "S2", "S5", "S7"]

    def _supplier_for(defects):  # noqa: F811
        if any(t in QUALITY_DEFECTS for t in (defects or [])):
            return "S3" if (abs(hash("|".join(defects))) % 2 == 0) else "S6"
        return GOOD_SUPPLIERS[abs(hash("|".join(defects or ["clean"]))) % len(GOOD_SUPPLIERS)]

    def _map_category(*hints):  # noqa: F811
        text = " ".join(str(h).lower() for h in hints if h)
        for cat, kws in CATEGORY_KEYWORDS:
            if any(k in text for k in kws):
                return cat
        return "饰品配件"

    def _clamp(x, lo, hi):  # noqa: F811
        return max(lo, min(hi, x))

    def _to_date(v):  # noqa: F811
        from datetime import date, datetime
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, str) and v.strip():
            s = v.strip().replace(" UTC", "").replace("Z", "")
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:  # noqa: BLE001
                    continue
        return None


# ---------------------------------------------------------------------------
# 解析入口
# ---------------------------------------------------------------------------
def parse_file(filename: str, content):
    """解析上传文件 → (cases, detected_type)。

    content：.csv 为 str（已解码文本）；.xlsx 为 bytes。
    detected_type ∈ {amazon, uci, thelook, rg, None}。
    """
    fn = (filename or "").lower()
    if fn.endswith(".csv") or (not fn.endswith(".xlsx") and isinstance(content, str)):
        cases, typ = parse_csv(content if isinstance(content, str) else content.decode("utf-8-sig"))
        return cases, typ
    if fn.endswith(".xlsx"):
        if openpyxl is None:
            raise RuntimeError("服务端未安装 openpyxl，无法解析 xlsx")
        return parse_xlsx(content)
    # 兜底：按内容猜测
    if isinstance(content, str):
        return parse_csv(content)
    return parse_xlsx(content)


# ---------------------------------------------------------------------------
# CSV 解析（TheLook 订单 / RG 格式 / Amazon csv）
# ---------------------------------------------------------------------------
def parse_csv(text: str):
    reader = csv.DictReader(io.StringIO(text))
    fields = [f.lower() for f in (reader.fieldnames or [])]
    fset = set(fields)
    # 类型识别
    if "returned_at" in fset and "order_id" in fset:
        return _parse_thelook_csv(reader), "thelook"
    if ("return_reason" in fset or "product_id" in fset) and "returned" in fset:
        return _parse_amazon_csv(reader), "amazon"
    if "sku" in fset:
        return _parse_rg_csv(reader), "rg"
    raise ValueError("无法识别的 CSV 格式（缺少 sku / order_id 等已知列）")


def _norm_header(h):
    return (h or "").strip().lower()


def _parse_thelook_csv(reader) -> list[dict]:
    out = []
    for i, row in enumerate(reader, 1):
        try:
            ret = (row.get("returned_at") or "").strip()
            if not ret:
                continue  # 仅退货行
            oid = str(row.get("order_id") or row.get("id") or i)
            pid = str(row.get("product_id") or "")
            cid = "TL-" + oid
            sku = ("TL-P" + pid) if pid else cid
            amt = _f(row.get("sale_price") or row.get("retail_price") or 0)
            name = (row.get("product_name") or row.get("name") or "").strip() or f"Product {pid}"
            cat = THELOOK_CAT_MAP.get((row.get("department") or row.get("category") or "").strip()) \
                or _map_category(name, row.get("category"), row.get("department"))
            defects = ["无明显瑕疵"]
            sup = _supplier_for(defects)
            out.append({
                "case_id": cid, "sku": sku, "sku_name": name, "category": cat,
                "supplier": sup, "supplier_name": SUPPLIERS.get(sup, sup),
                "platform": "TheLook", "language": "en", "region": "US",
                "amount": round(amt, 2), "date": _to_date(ret),
                "similarity": 0.85, "same_item": True, "defect_tags": defects,
                "defect_description": "退货（无原因字段）", "consistency": "待分析（无退货原因，无法判定责任归属）",
                "outcome": "待分析", "mode": "dataset",
            })
        except Exception as e:  # noqa: BLE001
            logger.warning("TheLook 行 %s 解析失败: %s", i, e)
    return out


def _parse_amazon_csv(reader) -> list[dict]:
    out = []
    for i, row in enumerate(reader, 1):
        try:
            if str(row.get("returned") or "").strip() not in ("1", "1.0", "True", "true"):
                continue
            reason = (row.get("return_reason") or "None")
            if reason in ("None", "nan", "", None):
                reason = "Changed mind"
            pid = str(row.get("product_id") or i)
            cat = (row.get("product_category") or "").strip()
            cid = "AMZ-" + pid
            amt = round(_f(row.get("price") or 0) * _f(row.get("quantity") or 1), 2)
            seller_rating = _f(row.get("seller_rating") or 4)
            return_score = _f(row.get("return_score") or 0.4)
            defects = [AMAZON_REASON_DEFECT.get(reason, "无明显瑕疵")]
            sim, same, outcome, consistency = _amazon_derive(reason, seller_rating, return_score, defects)
            out.append({
                "case_id": cid, "sku": cid, "sku_name": f"{cat} · {pid}", "category": AMAZON_CAT_MAP.get(cat, _map_category(cat)),
                "supplier": _supplier_for(defects), "supplier_name": SUPPLIERS.get(_supplier_for(defects), ""),
                "platform": "Amazon", "language": "en", "region": "US",
                "amount": amt, "date": _to_date(row.get("order_datetime") or row.get("date")),
                "similarity": sim, "same_item": same, "defect_tags": defects,
                "defect_description": reason, "consistency": consistency, "outcome": outcome, "mode": "dataset",
            })
        except Exception as e:  # noqa: BLE001
            logger.warning("Amazon CSV 行 %s 解析失败: %s", i, e)
    return out


def _parse_rg_csv(reader) -> list[dict]:
    """RG 格式 CSV（含 sku/amount 等；可有 case_id）。无 case_id 时用自然键生成稳定 id。"""
    out = []
    for i, row in enumerate(reader, 1):
        try:
            mapped = {}
            for h, v in row.items():
                if h is None or v in (None, ""):
                    continue
                mapped[_norm_header(h)] = str(v).strip()
            sku = mapped.get("sku")
            if not sku:
                continue
            cid = mapped.get("case_id") or ("RG-" + _stable_id(sku, mapped.get("date", ""), mapped.get("similarity", "")))
            amt = _f(mapped.get("amount", 0))
            sim = _f(mapped.get("similarity", 0.9))
            same = (mapped.get("same_item", "").lower() in ("true", "1", "是", "yes")) or sim >= SAME_ITEM_THRESHOLD
            defects = [t.strip() for t in re.split(r"[;,、]", mapped.get("defect_tags", "无明显瑕疵")) if t.strip()] or ["无明显瑕疵"]
            out.append({
                "case_id": cid, "sku": sku, "sku_name": mapped.get("sku_name", sku),
                "category": mapped.get("category", "饰品配件"),
                "supplier": mapped.get("supplier", _supplier_for(defects)),
                "supplier_name": mapped.get("supplier_name", SUPPLIERS.get(mapped.get("supplier", ""), "")),
                "platform": mapped.get("platform", "Amazon"), "language": mapped.get("language", "en"),
                "region": mapped.get("region", "US"), "amount": round(amt, 2),
                "date": _to_date(mapped.get("date")), "similarity": round(sim, 3), "same_item": same,
                "defect_tags": defects, "defect_description": mapped.get("defect_description", ""),
                "consistency": mapped.get("consistency", ""), "outcome": mapped.get("outcome", "待分析"),
                "mode": mapped.get("mode", "dataset"),
            })
        except Exception as e:  # noqa: BLE001
            logger.warning("RG CSV 行 %s 解析失败: %s", i, e)
    return out


# ---------------------------------------------------------------------------
# XLSX 解析（Amazon / UCI）
# ---------------------------------------------------------------------------
def parse_xlsx(content: bytes):
    bio = io.BytesIO(content)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    # 选表：优先 Sheet1(Amazon) / Online Retail(UCI) / 否则首表
    name = None
    for cand in ("Sheet1", "Online Retail"):
        if cand in wb.sheetnames:
            name = cand
            break
    if name is None:
        name = wb.sheetnames[0]
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], None
    hdr = [(_norm_header(str(h)) if h is not None else "") for h in rows[0]]
    hset = set(hdr)
    if "return_reason" in hset or ("product_id" in hset and "returned" in hset):
        return _parse_amazon_xlsx(rows, hdr), "amazon"
    if "invoiceno" in hset or "stockcode" in hset:
        return _parse_uci_xlsx(rows, hdr), "uci"
    raise ValueError("无法识别的 xlsx 格式（缺少 return_reason / InvoiceNo 等已知列）")


def _parse_amazon_xlsx(rows, hdr) -> list[dict]:
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        try:
            if str(r[idx.get("returned", -1)] if idx.get("returned", -1) >= 0 else "").strip() not in ("1", "1.0", "True", "true"):
                continue
            reason = str(r[idx["return_reason"]] or "None")
            if reason in ("None", "nan", ""):
                reason = "Changed mind"
            pid = str(r[idx["product_id"]])
            cat = str(r[idx.get("product_category", -1)] or "" if idx.get("product_category", -1) >= 0 else "")
            cid = "AMZ-" + pid
            amt = round(_f(r[idx["price"]]) * _f(r[idx.get("quantity", -1)] or 1), 2)
            seller_rating = _f(r[idx.get("seller_rating", -1)] or 4)
            return_score = _f(r[idx.get("return_score", -1)] or 0.4)
            defects = [AMAZON_REASON_DEFECT.get(reason, "无明显瑕疵")]
            sim, same, outcome, consistency = _amazon_derive(reason, seller_rating, return_score, defects)
            out.append({
                "case_id": cid, "sku": cid, "sku_name": f"{cat} · {pid}", "category": AMAZON_CAT_MAP.get(cat, _map_category(cat)),
                "supplier": _supplier_for(defects), "supplier_name": SUPPLIERS.get(_supplier_for(defects), ""),
                "platform": "Amazon", "language": "en", "region": "US",
                "amount": amt, "date": _to_date(r[idx.get("order_datetime", -1)] if idx.get("order_datetime", -1) >= 0 else None),
                "similarity": sim, "same_item": same, "defect_tags": defects,
                "defect_description": reason, "consistency": consistency, "outcome": outcome, "mode": "dataset",
            })
        except Exception as e:  # noqa: BLE001
            logger.warning("Amazon xlsx 行解析失败: %s", e)
    return out


def _parse_uci_xlsx(rows, hdr) -> list[dict]:
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        try:
            inv = r[idx.get("invoiceno", -1)] if idx.get("invoiceno", -1) >= 0 else None
            stock = r[idx.get("stockcode", -1)] if idx.get("stockcode", -1) >= 0 else None
            if inv is None or stock is None:
                continue
            qty = r[idx.get("quantity", -1)] if idx.get("quantity", -1) >= 0 else 0
            price = r[idx.get("unitprice", -1)] if idx.get("unitprice", -1) >= 0 else 0
            is_return = (isinstance(inv, str) and inv.startswith("C")) or (isinstance(qty, (int, float)) and qty < 0)
            if not is_return:
                continue
            amt = round(abs(_f(qty) * _f(price)), 2)
            if amt <= 0:
                continue
            cid = "UCI-" + str(inv)
            desc = str(r[idx.get("description", -1)] or "" if idx.get("description", -1) >= 0 else "")
            country = str(r[idx.get("country", -1)] or "GB" if idx.get("country", -1) >= 0 else "GB")
            cat = _map_category(desc)
            defects = ["无明显瑕疵"]
            sup = _supplier_for(defects)
            out.append({
                "case_id": cid, "sku": "UCI-" + str(stock), "sku_name": desc or str(stock),
                "category": cat, "supplier": sup, "supplier_name": SUPPLIERS.get(sup, sup),
                "platform": "UCI-Retail", "language": "en", "region": country,
                "amount": amt, "date": _to_date(r[idx.get("invoicedate", -1)] if idx.get("invoicedate", -1) >= 0 else None),
                "similarity": 0.85, "same_item": True, "defect_tags": defects,
                "defect_description": "退货（无原因字段）", "consistency": "待分析（无退货原因，无法判定责任归属）",
                "outcome": "待分析", "mode": "dataset",
            })
        except Exception as e:  # noqa: BLE001
            logger.warning("UCI xlsx 行解析失败: %s", e)
    return out


# ---------------------------------------------------------------------------
# 公共派生 + 工具
# ---------------------------------------------------------------------------
def _amazon_derive(reason, seller_rating, return_score, defects):
    """诚实派生 Amazon 案件的相似度/同款/胜负/一致性（确定性，无随机，重复导入可复现）。"""
    if reason in REMORSE_REASONS:
        base = 0.91
    elif reason in ("Wrong item", "Not as described"):
        base = 0.64
    elif reason == "Size/fit issue":
        base = 0.85
    else:
        base = 0.88
    sim = round(_clamp(base + (return_score - 0.5) * 0.1, 0.5, 0.99), 3)
    same = sim >= SAME_ITEM_THRESHOLD
    wp = 0.72 if reason in REMORSE_REASONS else 0.22
    wp += (seller_rating - 4) * 0.05
    wp -= (return_score - 0.4) * 0.1
    wp = _clamp(wp, 0.03, 0.95)
    # 确定性胜负：按比例切分（<wp 赢，<wp+0.22 部分退款，否则输）
    outcome = "赢" if wp >= 0.78 else ("部分退款" if wp >= 0.5 else "输")
    consistency = ("一致（疑似非质量原因，倾向买家责任）"
                   if (same and defects == ["无明显瑕疵"])
                   else "存在差异（货不对板 / 运输或质量瑕疵）")
    return sim, same, outcome, consistency


def _f(v) -> float:
    try:
        return float(v)
    except Exception:  # noqa: BLE001
        return 0.0


def _stable_id(*parts) -> str:
    import hashlib
    h = hashlib.md5("|".join(str(p) for p in parts).encode("utf-8")).hexdigest()[:8].upper()
    return h


if __name__ == "__main__":
    import sys
    # 简单自测：python dataset_parse.py <file>
    if len(sys.argv) > 1:
        with open(sys.argv[1], "rb") as f:
            data = f.read()
        cases, typ = parse_file(sys.argv[1], data if sys.argv[1].lower().endswith(".xlsx") else data.decode("utf-8-sig", errors="replace"))
        print("detected:", typ, "cases:", len(cases))
        for c in cases[:3]:
            print(" ", c["case_id"], c["platform"], c["category"], c["amount"], c["date"], c["outcome"])
